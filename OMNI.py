import pandas as pd
import numpy as np
from openpyxl import load_workbook
from openpyxl.styles import NamedStyle
from openpyxl.utils import get_column_letter

# ------------------------------
import os
import sys
# ------------------------------
import time
from datetime import datetime
import re
# ------------------------------

user_id = os.getlogin()
if getattr(sys, 'frozen', False):

    folder_path = os.path.dirname(sys.executable)

else:

    folder_path = os.path.dirname(os.path.abspath(__file__))

xl_erisite_name = 'base_edb.xlsx'
xl_omni_name = 'base_omni.xlsx'
xl_fc_name = 'de-para_colunas.xlsx'
xl_vendor_name = 'colunas_dependentes.xlsx'

xl_erisite_file = os.path.join(folder_path, xl_erisite_name)
xl_omni_file = os.path.join(folder_path, xl_omni_name)
xl_fc_file = os.path.join(folder_path, xl_fc_name)
xl_vendor_file = os.path.join(folder_path, xl_vendor_name)

def user_menu():
    print('-' * 90)
    print('OMNI Data Converter'.center(90))
    print('-' * 90)
    
    user_option = input('\nDeseja iniciar a geração de um arquivo OMNI? [S/N]:').upper()
    return user_option

def data_update_file_creation(xl_erisite, xl_omni, xl_fc, xl_vendor):

    # Verifica se os arquivos estão contidos na pasta ou se estão corrompidos
    try:
        df_fc = pd.read_excel(xl_fc, sheet_name=0)
        fields_correlation = dict(zip(df_fc['Erisite'], df_fc['OMNI']))

    except ValueError:
        print(f'O arquivo "{xl_fc}" apresenta problemas ou não foi encontrado. Por favor, corrigir o problema antes de tentar novamente.')
        time.sleep(3)
        user_menu()

    try:
        df_vendor = pd.read_excel(xl_vendor, sheet_name=0)
        vendor_analysis = []
        data_type_list = []

        for _, row in df_vendor.iterrows():
            vendor_analysis.append(list([row['OMNI Column'], row['Related Vendor']]))
            data_type_list.append(row['Data Type'])

    except ValueError:
        print(f'O arquivo "{xl_vendor}" apresenta problemas ou não foi encontrado. Por favor, corrigir o problema antes de tentar novamente.')
        time.sleep(3)
        user_menu()

    try:
        df_erisite = pd.read_excel(xl_erisite, sheet_name=0, keep_default_na=False)
        df_corr_erisite = df_erisite.rename(columns=fields_correlation)
        df_corr_erisite = df_corr_erisite.set_index('UID_IDPMTS')

    except ValueError:
        print(f'O arquivo "{xl_erisite}" apresenta problemas ou não foi encontrado. Por favor, corrigir o problema antes de tentar novamente.')
        time.sleep(3)
        user_menu()

    try:
        df_omni = pd.read_excel(xl_omni, sheet_name=0, keep_default_na=False)
        df_omni = df_omni.set_index('UID_IDPMTS')

    except ValueError:
        print(f'O arquivo "{xl_omni}" apresenta problemas ou não foi encontrado. Por favor, corrigir o problema antes de tentar novamente.')
        time.sleep(3)
        user_menu()

    changes_log = []
    num_change = 0

    df_change = df_omni.reset_index().copy()
    df_change = df_change.rename(columns={"index": "UID_IDPMTS"})

    date_columns = []

    prev_ID = ''

    # Verifica IDs duplicados na base Erisite

    if df_corr_erisite.index.duplicated().any():
                                    
        duplicated_ID_erisite = df_corr_erisite.index[df_corr_erisite.index.duplicated()].unique()
        print('O arquivo do Erisite possui IDs duplicados:')
        print(duplicated_ID_erisite)
        print('\nPor favor, corrija os itens acima primeiro antes de rodar novamente.')
        print('-' * 90)

    else:
        
        analysis_sample = len(df_corr_erisite.columns) * len(df_corr_erisite)

        for ids, (r, val) in enumerate(df_corr_erisite.iterrows()):

            try:

                omni_x = df_omni.index.get_loc(r)

            except KeyError:

                changes_log.append({"ID PMTS": r, "Coluna Atualizada": '', "Valor Anterior": '', "Valor Novo": '', "Status ID": 'Novo'})

        for j in range(0, len(df_corr_erisite.columns)):

            curr_data_type = ''

            for i in range(0, len(vendor_analysis)):

                # Verifica se a coluna está na relação de campos com os vendors
                if vendor_analysis[i][0] == df_corr_erisite.columns[j]:
                    
                    for linha, (v, curr_omni_val) in enumerate(df_corr_erisite.iterrows()):
                        
                        progress = ((linha + len(df_corr_erisite) * (j)) / analysis_sample) * 100

                        print(f'Progresso: {progress:.2f}% ...', end='\r')

                        erisite_x = df_corr_erisite.index.get_loc(v)
                        erisite_y = j

                        data_type = data_type_list[i]
                        curr_data_type = data_type

                        new_val = formatted_val(df_corr_erisite.iloc[erisite_x, erisite_y], data_type)
                        
                        try:
                            omni_x = df_omni.index.get_loc(v)
                            omni_y = df_omni.columns.get_loc(vendor_analysis[i][0])

                            vendor_col = df_omni.columns.get_loc(vendor_analysis[i][1])

                            if df_omni.iloc[omni_x, vendor_col] == 'Ericsson':

                                curr_val = formatted_val(df_omni.iloc[omni_x, omni_y], data_type)

                                #print(f'ID EDB: {df_corr_erisite.index.get_loc(v)} - Campo EDB: {df_corr_erisite.columns[j]} - Valor Atual: {df_corr_erisite.iloc[erisite_x, erisite_y]} / ID OMNI: {df_omni.index.get_loc(v)} - Campo OMNI: {vendor_analysis[i][0]} - Valor Anterior: {df_omni.iloc[omni_x, omni_y]}')

                                if new_val != None and new_val != curr_val:

                                    df_change.iloc[omni_x, omni_y + 1] = new_val
                                    changes_log.append({"ID PMTS": v, "Coluna Atualizada": vendor_analysis[i][0], "Valor Anterior": curr_val, "Valor Novo": new_val, "Status ID": 'Existente'})
                                    num_change += 1

                                elif new_val == curr_val:

                                    df_change.iloc[omni_x, omni_y + 1] = curr_val

                        except KeyError:
                            pass

                        except ValueError:
                            pass
                        

                # Formata as colunas de data presentes no fields correlation de vendors
            if curr_data_type == 'Date':
                
                curr_col = df_omni.columns[omni_y + 1]

                date_columns.append(curr_col)

    if num_change == 0:

        print('Não houveram alterações no arquivo.')
        print('-' * 90)
        print('\n')

    else:
        
        today_date = datetime.today()
        today_text = datetime.strftime(today_date, '%Y-%m-%d_%H-%M-%S')

        upload_file_name = f'OMNI - Upload - {today_text}.xlsx'
        xl_upload_file = os.path.join(folder_path, upload_file_name)

        log_name = f'OMNI - Log - {today_text}.xlsx'
        xl_log_file = os.path.join(folder_path, log_name)
        df_log = pd.DataFrame(changes_log)
        
        
        # Geração dos arquivos

        df_change.to_excel(xl_upload_file, sheet_name='Upload',index=False)
        wb = load_workbook(xl_upload_file)
        wsht = wb['Upload']

        data_style_format = NamedStyle(name='data_style', number_format='DD/MM/YYYY')

        for col in date_columns:
           
           col_idy = df_change.columns.get_loc(col)
           date_col = get_column_letter(col_idy)
            
           for cell in wsht[date_col]:
               
               cell.style = data_style_format

        wb.save(xl_upload_file)

        df_log.to_excel(xl_log_file, index=False)

        print('\nArquivo OMNI para upload gerado com sucesso!')
        print(f'Arquivo de log com as {num_change} atualizações gerado com sucesso!')
        print('-' * 90)

# Função para formatar valor de acordo com o Fields Correlation
def formatted_val(value, data_type):
    
    if value is None or value == '' or value == 'NaT' or value == 'NaN':

        return None
                
    form_val = str(value).strip()

    if data_type == 'Text':

        return form_val
    
    if form_val.isnumeric():
            
        return int(form_val)
    
    if data_type != 'Date':

        return value

    #'%Y/%m/%d %H:%M:%S'
    date_formats = ['%d/%m/%Y']
    invalid_fmt = True

    regex_ddmmyyyy = r'^(0[1-9]|[12][0-9]|3[01])/(0[1-9]|1[0-2])/\d{4}$'
    if re.match(regex_ddmmyyyy, str(value)):
        try:

            date_form_val = datetime.strptime(str(value), '%d/%m/%Y')
            #date_form_val_corr = date_form_val.strftime("%d/%m/%Y")

            invalid_fmt = False
            #print(f'O valor atual é: {date_form_val}.')
            return date_form_val
    
        except ValueError:
            return None

    for fmt in date_formats:
        try:
            date_form_val = datetime.strptime(str(value), fmt)
            date_form_val = date_form_val.strftime('%d/%m/%Y')
            invalid_fmt = False

            return date_form_val

        except ValueError:
            continue

    if invalid_fmt:
        return value

    if NaN_values(value) or NaT_values(value):
        return None     

def NaN_values(value):
    try:
        if value == 'NA' or value == 'NOK':
            return value
        else:
            return np.isnan(value)
    
    except (TypeError, ValueError):
        return False
    
def NaT_values(value):
    try:
        return np.isnat(value)
    
    except (TypeError, ValueError):
        return False

def program():
    while True:

        user_option = user_menu()

        if user_option == 'S':
            data_update_file_creation(xl_erisite_file, xl_omni_file, xl_fc_file, xl_vendor_file)
            print('\nObrigado pelo aguardo!'.center(90, ' '))
            time.sleep(5)
            break

        elif user_option == 'N':
            print('\nFinalizando o programa...')
            time.sleep(5)
            break

        else:
            print('Comando inválido. Por favor, escolha entre as opções [S/N] solicitadas.')
            user_menu()

program()